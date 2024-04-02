import pandas as pd
from rapidfuzz import fuzz
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import re
from io import BytesIO
from rapidfuzz import process



# Function to clean column names by removing leading and trailing spaces
def clean_column_names(df):
    df.columns = df.columns.str.strip()
    return df


# Function to standardize common street suffix abbreviations
def standardize_address(address):
    replacements = {
        r'\bSt\b': 'street',
        r'\bRd\b': 'road',
        r'\bAve\b': 'avenue',
        r'\bPl\b': 'place',
        r'\bDr\b': 'drive',
        r'\bLn\b': 'lane',
        r'\bBlvd\b': 'boulevard',
        r'\bCt\b': 'court',
        r'\bAlly\b': 'alley',
        r'\bAlwy\b': 'alleyway',
        r'\bArc\b': 'arcade',
        r'\bBasn\b': 'basin',
        r'\bBch\b': 'beach',
        r'\bBend\b': 'bend',
        r'\bBlk\b': 'block',
        r'\bBvd\b': 'boulevard',
        r'\bBdge\b': 'bridge',
        r'\bBdwy\b': 'broadway',
        r'\bBypa\b': 'bypass',
        r'\bBywy\b': 'byway',
        r'\bCaus\b': 'causeway',
        r'\bCn\b': 'central',
        r'\bCtr\b': 'centre',
        r'\bCnwy\b': 'centreway',
        r'\bCh\b': 'chase',
        r'\bCir\b': 'circle',
        r'\bCct\b': 'circuit',
        r'\bCl\b': 'close',
        r'\bCon\b': 'concourse',
        r'\bCnr\b': 'corner',
        # Postal delivery type abbreviations
        r'\bcare po\b': 'care Of Post Office',
        r'\bcma\b': 'community Mail Agent',
        r'\bcmb\b': 'community Mail Bag',
        r'\bgpo box\b': 'general Post Office Box',
        r'\blocked bag\b': 'locked Mail Bag Service',
        r'\bms\b': 'mail Service',
        r'\bpo box\b': 'post Office Box',
        r'\bprivate bag\b': 'private mail bag service',
        r'\brsd\b': 'roadside delivery',
        r'\brmb\b': 'roadside mail bag',
        r'\brms\b': 'roadside mail service',
        r'\bcpa\b': 'community postal agent',
        r'\bstrp\b': 'strip',
        r'\bsbwy\b': 'subway',
        r'\bthor\b': 'thoroughfare',
        r'\btlwy\b': 'tollway',
        r'\btwrs\b': 'towers',
        r'\btrk\b': 'track',
        r'\btrlr\b': 'trailer',
        r'\btri\b': 'triangle',
        r'\btkwy\b': 'trunkway',
        r'\bturn\b': 'turn',
        r'\bupas\b': 'underpass',
        r'\bup\b': 'upper',
        r'\bvale\b': 'vale',
        r'\bvdct\b': 'viaduct',
        r'\bvlls\b': 'villas',
        r'\bvsta\b': 'vista',
        r'\bwalk\b': 'walk',
        r'\bwkyw\b': 'walkway',
        r'\bw\b': 'west',
        r'\bwhrf\b': 'wharf',
        r'\bwynd\b': 'wynd',
        r'\byard\b': 'yard',
        r'\brch\b': 'reach',
        r'\bres\b': 'reserve',
        r'\brtt\b': 'retreat',
        r'\brgwy\b': 'ridgeway',
        r'\browy\b': 'right of Way',
        r'\brvr\b': 'river',
        r'\brvwy\b': 'riverway',
        r'\brvra\b': 'riviera',
        r'\brds\b': 'roads',
        r'\brdwy\b': 'roadway',
        r'\brnde\b': 'ronde',
        r'\brsbl\b': 'rosebowl',
        r'\brty\b': 'rotary',
        r'\brnd\b': 'round',
        r'\brte\b': 'route',
        r'\brun\b': 'run',
        r'\bswy\b': 'service way',
        r'\bsdng\b': 'siding',
        r'\bslpe\b': 'slope',
        r'\bsnd\b': 'sound'

    }

    # Use regex to perform replacements only where the whole word matches
    for old, new in replacements.items():
        address = re.sub(old, new, address, flags=re.IGNORECASE)

    return address


def combined_matching(target_address, target_name, lookout_dataset):
    best_address = 'none'
    best_name_match_for_best_address = 'none'
    best_mobile_number_for_best_address = 'none'  # Add mobile number
    highest_address_score = 0
    highest_name_score_for_best_address = 0  # Score for the name corresponding to the best address match

    # First, find the best address match
    for _, row in lookout_dataset.itertuples():
        address_score = fuzz.ratio(target_address, row['Address']) / 100

        if address_score > highest_address_score:
            highest_address_score = address_score
            best_address = row['Address']
            best_name_match_for_best_address = row['Full Name']
            highest_name_score_for_best_address = fuzz.ratio(target_name, row['Last Name']) / 100
            best_mobile_number_for_best_address = row['Mobile']  # Capture the mobile number

            if address_score*highest_name_score_for_best_address > 0.7:
                break

    combined_score = highest_address_score * highest_name_score_for_best_address

    return best_address, highest_address_score, best_name_match_for_best_address, highest_name_score_for_best_address, combined_score, best_mobile_number_for_best_address

def confidence(combined_score):
    if combined_score >= 0.70:
        return 'High'
    elif combined_score >= 0.40:
        return 'Medium'
    else:
        return 'Low'


def normalize_combined_score(combined_score):
    return (combined_score - combined_score.min()) / (combined_score.max() - combined_score.min())


# Function to save DataFrame to an Excel file in-memory and return it
def create_styled_excel(df, confidence_col_index):
    # Create a BytesIO buffer to write the Excel file to
    output = BytesIO()
    # Write the DataFrame to an Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Apply styles after writing
        for row in range(2, worksheet.max_row + 1):  # Start from 2 to skip the header
            cell = worksheet.cell(row=row, column=confidence_col_index)
            value = cell.value
            color = 'FFFFFF'  # Default white
            if value == 'High':
                color = '90EE90'  # Light green
            elif value == 'Medium':
                color = 'FFD700'  # Gold
            elif value == 'Low':
                color = 'FF6347'  # Tomato red
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Rewind the buffer
    output.seek(0)
    return output



# Precompute the scoring matrix for the lookout_dataset
def precompute_scores(target_address, target_name, lookout_dataset):
    # Assuming lookout_dataset is a DataFrame with 'Address' and 'Full Name'
    # Convert the address and name columns to lists for processing.
    addresses = lookout_dataset['Address'].tolist()
    names = lookout_dataset['Full Name'].tolist()

    # Precompute the scores as dictionaries if the list of unique values is smaller than the dataset
    address_scores = {address: fuzz.ratio(target_address, address) for address in set(addresses)}
    name_scores = {name: fuzz.ratio(target_name, name) for name in set(names)}

    return address_scores, name_scores


# Optimized combined_matching function
def optimized_combined_matching(target_address, target_name, address_scores, name_scores, lookout_dataset):
    best_address = max(address_scores, key=address_scores.get)
    highest_address_score = address_scores[best_address]

    # Extract best name match for the best address
    corresponding_name = lookout_dataset[lookout_dataset['Address'] == best_address]['Full Name'].iloc[0]
    highest_name_score = name_scores[corresponding_name]

    # Get mobile number for the best address match
    best_mobile_number_for_best_address = lookout_dataset[lookout_dataset['Address'] == best_address]['Mobile'].iloc[0]

    combined_score = highest_address_score * highest_name_score

    return best_address, highest_address_score, corresponding_name, highest_name_score, combined_score, best_mobile_number_for_best_address




# Assuming dataset1 is your lookup dataset and dataset2 is your target dataset.

def find_best_matches(target_dataset, lookup_dataset):
    results = []

    # Preparing the lookup lists from the lookup dataset
    lookup_addresses = lookup_dataset['Address'].tolist()
    lookup_names = lookup_dataset['Full Name'].tolist()  # Ensure this is correct based on your dataset structure

    for index, row in target_dataset.iterrows():
        target_address = row['Address']
        target_name = row["Owner's Name"]

        # Find the best match for the address in lookup_dataset
        best_address_match = process.extractOne(target_address, lookup_addresses, scorer=fuzz.ratio)

        if best_address_match:
            # Retrieve the best address match details
            best_address_score = best_address_match[1]
            best_address_index = best_address_match[2]
            best_matched_address = lookup_dataset.iloc[best_address_index]

            # Now match the owner's name only within the context of the matched address
            # Assuming the best name should be matched within the same record as the best address
            best_name_score = fuzz.ratio(target_name, best_matched_address['Last Name'])

            results.append({
                'Best Match Address': best_matched_address['Address'],
                'Combined Score': float(best_address_score) * float(best_name_score)/10000,
                'Best Match Name': best_matched_address['Full Name'],
                'Mobile': best_matched_address['Mobile'],
            })
        else:
            # Append none or default values if no address match is found
            results.append({
                'Best Match Address': 'none',
                'Combined Score': 0,
                'Best Match Name': 'none',
                'Mobile': 'none',
            })

    return pd.DataFrame(results)

# Ensure that dataset1 and dataset2 are defined and structured correctly before calling this function

# Example usage:
#matches_df = find_best_matches(dataset2, dataset1)

# Now 'matches_df' contains the best matches from the lookup dataset for each entry in the target dataset.
